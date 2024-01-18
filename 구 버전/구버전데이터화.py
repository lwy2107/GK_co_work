import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle 

def process_file(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # A1부터 N6까지의 범위 선택하여 데이터 삭제
    merged_ranges = sheet.merged_cells.ranges
    for row in range(1, 1000):  # 1부터 6까지의 행
        for col in range(1, 15):  # 1부터 14까지의 열 (A부터 N까지)
            cell = sheet.cell(row=row, column=col)
            for merged_range in merged_ranges:
                if cell.coordinate in merged_range:
                    # 병합된 셀 범위에 포함된 경우
                    sheet.unmerge_cells(merged_range.coord)
                    break  # 해당 셀이 병합된 범위에 속하면 루프 종료
    sheet.delete_rows(1, 6)  # 1부터 6번째 행까지 삭제

    new_style = NamedStyle(name='custom_style', number_format='HH:mm')

    # 첫 번째 열에 첫 번째 스타일 적용
    column_number = 5
    for cell in sheet.iter_cols(min_col=column_number, max_col=column_number):
        for col in cell:
            col.style = new_style

    # D열 숫자 + E열 문자 + F열 숫자 조합하여 하나의 문자열로 만들기
    combined_values = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=4, max_col=6):
        d_value = str(row[0].value) if row[0].value is not None else ""  # D열의 숫자 값을 문자열로 변환
        e_value = str(row[1].value) if row[1].value is not None else ""  # E열의 문자 값을 문자열로 변환
        f_value = str(row[2].value) if row[2].value is not None else ""  # F열의 숫자 값을 문자열로 변환
        combined_value = d_value + e_value + f_value  # 조합된 문자열 생성
        combined_values.append([combined_value])

    # 조합된 값 열에 추가
    for idx, value in enumerate(combined_values, start=1):
        sheet.cell(row=idx, column=4).value = value[0]

    combined_values = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=7, max_col=9):
        d_value = str(row[0].value) if row[0].value is not None else ""  # D열의 숫자 값을 문자열로 변환
        e_value = str(row[1].value) if row[1].value is not None else ""  # E열의 문자 값을 문자열로 변환
        f_value = str(row[2].value) if row[2].value is not None else ""  # F열의 숫자 값을 문자열로 변환
        combined_value = d_value + e_value + f_value  # 조합된 문자열 생성
        combined_values.append([combined_value])

    # 조합된 값 열에 추가
    for idx, value in enumerate(combined_values, start=1):
        sheet.cell(row=idx, column=7).value = value[0]

    columns_to_clear = [5, 6, 8, 9]  # E, F, H, I열에 해당하는 열 번호
    for col in columns_to_clear:
        for cell in sheet.iter_cols(min_col=col, max_col=col):
            for row in cell:
                row.value = None  # 값 삭제

    for source, destination in zip(sheet.iter_cols(min_col=7, max_col=7, min_row=1), sheet.iter_cols(min_col=5, max_col=5, min_row=1)):
        for source_cell, dest_cell in zip(source, destination):
            dest_cell.value = source_cell.value  # G열의 값을 E열로 복사
            source_cell.value = None  # G열의 값을 지움

    # J~N열의 내용을 F~J열로 이동
    for source_col, destination_col in zip(range(10, 15), range(6, 11)):
        for source, destination in zip(sheet.iter_cols(min_col=source_col, max_col=source_col, min_row=1), sheet.iter_cols(min_col=destination_col, max_col=destination_col, min_row=1)):
            for source_cell, dest_cell in zip(source, destination):
                dest_cell.value = source_cell.value  # J~N열의 값을 F~J열로 복사
                source_cell.value = None  # J~N열의 값을 지움

    # 변경된 내용 저장
    workbook.save(file_path)
    print(f"변경된 파일: {file_path}")

def select_file():
    root = tk.Tk()
    root.withdraw()  # 기본 창 숨기기
    file_path = filedialog.askopenfilename(title="파일을 선택하세요", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        process_file(file_path)
    else:
        print("파일 선택이 취소되었습니다.")

if __name__ == "__main__":
    select_file()




