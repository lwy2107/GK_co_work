import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

def process_folder(folder_path):
    # 폴더 내의 모든 파일에 대한 처리
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            process_file(file_path)

def process_file(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # A1부터 N6까지의 범위 선택하여 데이터 삭제
    merged_ranges = sheet.merged_cells.ranges
    for row in range(1, 1000):  # 1부터 6까지의 행
        for col in range(1, 20):  # 1부터 14까지의 열 (A부터 Q까지)
            cell = sheet.cell(row=row, column=col)
            for merged_range in merged_ranges:
                if cell.coordinate in merged_range:
                    # 병합된 셀 범위에 포함된 경우
                    sheet.unmerge_cells(merged_range.coord)
                    break  # 해당 셀이 병합된 범위에 속하면 루프 종료


    new_style = NamedStyle(name='custom_style', number_format='HH:mm')

    # 첫 번째 열에 첫 번째 스타일 적용
    column_number = 5
    for cell in sheet.iter_cols(min_col=column_number, max_col=column_number):
        for col in cell:
            col.style = new_style

    # D열 숫자 + E열 문자 + F열 숫자 조합하여 하나의 문자열로 만들기
    combined_values = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=3, max_col=5):
        d_value = str(row[0].value) if row[0].value is not None else ""  # D열의 숫자 값을 문자열로 변환
        e_value = str(row[1].value) if row[1].value is not None else ""  # E열의 문자 값을 문자열로 변환
        f_value = str(row[2].value) if row[2].value is not None else ""  # F열의 숫자 값을 문자열로 변환
        combined_value = d_value + e_value + f_value  # 조합된 문자열 생성
        combined_values.append([combined_value])

    # 조합된 값 열에 추가
    for idx, value in enumerate(combined_values, start=1):
        sheet.cell(row=idx, column=3).value = value[0]

    combined_values = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=6, max_col=8):
        d_value = str(row[0].value) if row[0].value is not None else ""  # D열의 숫자 값을 문자열로 변환
        e_value = str(row[1].value) if row[1].value is not None else ""  # E열의 문자 값을 문자열로 변환
        f_value = str(row[2].value) if row[2].value is not None else ""  # F열의 숫자 값을 문자열로 변환
        combined_value = d_value + e_value + f_value  # 조합된 문자열 생성
        combined_values.append([combined_value])

    # 조합된 값 열에 추가
    for idx, value in enumerate(combined_values, start=1):
        sheet.cell(row=idx, column=6).value = value[0]



    # G열의 값을 E열로 복사
    for source_col, destination_col in zip(range(6, 7), range(4, 5)):
        for source, destination in zip(sheet.iter_cols(min_col=source_col, max_col=source_col, min_row=1),
                                        sheet.iter_cols(min_col=destination_col, max_col=destination_col, min_row=1)):
            for source_cell, dest_cell in zip(source, destination):
                if source_cell.value is not None:
                    dest_cell.value = source_cell.value # G열의 값이 있으면 E열로 복사

    # J~Q열의 내용을 F~M열로 이동
    for source_col, destination_col in zip(range(9, 11), range(5, 7)):
        for source, destination in zip(sheet.iter_cols(min_col=source_col, max_col=source_col, min_row=1),
                                    sheet.iter_cols(min_col=destination_col, max_col=destination_col, min_row=1)):
            for source_cell, dest_cell in zip(source, destination):
                dest_cell.value = source_cell.value  # J~Q열의 값을 F~M열로 복사
                source_cell.value = None  # J~Q열의 값을 지움

    # H, I, J, O열의 2행을 한 칸 위로 이동
    for col in [8, 9, 10, 11]:  # H, I, J, O 열에 해당하는 열 번호
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row - 1, column=col).value = sheet.cell(row=row, column=col).value
            sheet.cell(row=row, column=col).value = None  # 이동된 칸의 값을 지움
    # H, I, J, O열의 1행 이후의 값을 한 칸 아래로 이동
    for col in [8, 9, 10, 11]:  # H, I, J, O 열에 해당하는 열 번호
        for row in range(sheet.max_row, 1, -1):
            sheet.cell(row=row + 1, column=col).value = sheet.cell(row=row, column=col).value
    for col in [8, 9, 10, 11]:  # H, I, J, O 열에 해당하는 열 번호
        sheet.cell(row=2, column=col).value = None
    # 변경된 내용 저장
    workbook.save(file_path)

def select_folder():
    root = tk.Tk()
    root.withdraw()  # 기본 창 숨기기
    folder_path = filedialog.askdirectory(title="폴더를 선택하세요")
    if folder_path:
        process_folder(folder_path)
    else:
        print("폴더 선택이 취소되었습니다.")

if __name__ == "__main__":
    select_folder()